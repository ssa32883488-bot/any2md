"""Fast CPU paths: digital PDF, Office, without Paddle GPU."""

from __future__ import annotations

import time
from pathlib import Path

from file_types import is_office, suffix
from docx_structured import docx_to_markdown_body
from output_layout import OutputBatch
from pdf_structured import pdf_needs_rich_layout, pdf_to_markdown_structured


def _log(fn, stage: str, message: str, **extra) -> None:
    fn(stage, message, **extra)


def pdf_to_markdown(path: Path, batch: OutputBatch, log_event) -> Path:
    """Structured CPU extraction: headings, tables, inline images."""
    t0 = time.perf_counter()
    rich = pdf_needs_rich_layout(path)
    if rich:
        log_event(
            "init",
            "检测到表格/图片，使用结构化提取（保留版式，仍走 CPU 快路径）",
            route="text",
        )
    out = pdf_to_markdown_structured(
        path,
        batch.md_path(path.stem),
        batch.assets_dir(path.stem),
        log_event,
    )
    log_event(
        "save",
        f"已保存 {out.name}（结构化快路径，{time.perf_counter() - t0:.1f}s）",
        route="text",
        elapsed_s=round(time.perf_counter() - t0, 1),
    )
    return out


def docx_to_markdown(path: Path, batch: OutputBatch, log_event) -> Path:
    from docx import Document

    t0 = time.perf_counter()
    doc = Document(path)
    body = docx_to_markdown_body(doc, title=path.stem)
    out = batch.md_path(path.stem)
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(body, encoding="utf-8")
    log_event(
        "save",
        f"已保存 {out.name}（Word 结构化转换，{time.perf_counter() - t0:.1f}s）",
        route="text",
        elapsed_s=round(time.perf_counter() - t0, 1),
    )
    return out


def xlsx_to_markdown(path: Path, batch: OutputBatch, log_event) -> Path:
    from openpyxl import load_workbook

    t0 = time.perf_counter()
    wb = load_workbook(path, read_only=True, data_only=True)
    parts = [f"# {path.stem}\n"]
    for sheet in wb.worksheets:
        parts.append(f"## {sheet.title}\n")
        rows_out: list[str] = []
        for row in sheet.iter_rows(values_only=True):
            cells = ["" if v is None else str(v).replace("\n", " ").strip() for v in row]
            if not any(cells):
                continue
            rows_out.append("| " + " | ".join(cells) + " |")
        if rows_out:
            parts.extend(rows_out)
            parts.append("")
    wb.close()

    out = batch.md_path(path.stem)
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text("\n".join(parts) + "\n", encoding="utf-8")
    log_event(
        "save",
        f"已保存 {out.name}（Excel 转换，{time.perf_counter() - t0:.1f}s）",
        route="text",
        elapsed_s=round(time.perf_counter() - t0, 1),
    )
    return out


def convert_fast(path: Path, batch: OutputBatch, log_event) -> Path:
    """Convert supported office/digital PDF via CPU-only fast path."""
    ext = suffix(path)
    log_event("init", f"使用 CPU 快路径处理 {path.name}", route="text", file=path.name)
    if ext == ".pdf":
        return pdf_to_markdown(path, batch, log_event)
    if ext == ".docx":
        return docx_to_markdown(path, batch, log_event)
    if ext in (".xlsx", ".xls"):
        if ext == ".xls":
            raise RuntimeError("旧版 .xls 请另存为 .xlsx 后重试")
        return xlsx_to_markdown(path, batch, log_event)
    if ext == ".doc":
        raise RuntimeError("旧版 .doc 请在 Word 中另存为 .docx 后重试")
    if is_office(path):
        raise RuntimeError(f"不支持的 Office 格式：{ext}")
    raise RuntimeError(f"快路径不支持：{ext}")


def can_fast_convert(path: Path) -> bool:
    ext = suffix(path)
    if ext == ".pdf":
        try:
            import fitz  # noqa: F401
        except ImportError:
            return False
        return True
    if ext == ".docx":
        try:
            import docx  # noqa: F401
        except ImportError:
            return False
        return True
    if ext == ".xlsx":
        try:
            import openpyxl  # noqa: F401
        except ImportError:
            return False
        return True
    return False
